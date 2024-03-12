import re
from PyPDF2 import PdfReader
import pandas as pd
from openpyxl import load_workbook
import glob

def extraer_y_limpiar_texto_de_pdfs(ruta_directorio):
    # Buscar todos los archivos PDF en el directorio dado
    archivos_pdf = glob.glob(f"{ruta_directorio}/*.pdf")
    textos_extraidos = []

    for ruta_archivo in archivos_pdf:
        with open(ruta_archivo, 'rb') as archivo:
            lector_pdf = PdfReader(archivo)
            texto_total = ""
            for pagina in range(len(lector_pdf.pages)):
                texto_pagina = lector_pdf.pages[pagina].extract_text()
                if texto_pagina:
                    texto_limpio = re.sub(r'\s+', ' ', texto_pagina)
                    texto_total += texto_limpio
            textos_extraidos.append(texto_total)
    
    return textos_extraidos

def extraer_informacion(texto):
    patrones = {
        "NIT": r"(\d{9}-\d{1})", 
        "Razón social": r"(?:NOMBRE|RAZÓN SOCIAL|Razón social|Nombre):\s*([\s\S]*?)(?=\n\S)",  
        "Correo Electrónico": r"[\w.-]+@[\w.-]+.\w+",  
        "Teléfono": r"(\b\d{10}\b|\b\d{7}\b)",  
        "Municipio": r"(?:MUNICIPIO|DOMICILIO|ADMINISTRACIÓN DIAN|Municipio):\s*([\w\s]+)",  
        "Código CIIU": r"(?:Actividad principal Código CIIU|ACTIVIDAD PRINCIPAL|Código ciiu|Código CIIU)\s*:\s*([A-Za-z]?\d{4})",  
    }
    resultados = {}
    for clave, patron in patrones.items():
        coincidencias = re.findall(patron, texto)
        if coincidencias:
            valor = coincidencias[0]
            resultados[clave] = valor.strip() if clave != "Razón social" else valor
        else:
            resultados[clave] = "No encontrado"

    return resultados

texto_pdf = extraer_y_limpiar_texto_de_pdfs(r"D:\Tesis maestria\Langchain\pdfs")
informacion = extraer_informacion(texto_pdf)

for clave, valor in informacion.items():
    print(f"{clave}: {valor}")
df = pd.DataFrame([informacion])
# Ruta del archivo Excel existente
ruta_excel = r'D:\Tesis maestria\Data_Tesis_Maestria\PAF BBVA_CLIENTE_MONTO_FECHA (V2024-0)_GENERICO.xlsx'

# Cargar el libro de trabajo existente
book = load_workbook(ruta_excel)

# Añade una nueva hoja llamada "Data_py" o selecciona si ya existe
if 'Data_py' not in book.sheetnames:
    book.create_sheet('Data_py')
sheet = book['Data_py']

# Escribe los encabezados (claves)
for j, header in enumerate(df.columns):
    sheet.cell(row=1, column=j + 1, value=header)

# Escribe los valores del DataFrame
for i, row in enumerate(df.values):
    for j, value in enumerate(row):
        sheet.cell(row=i + 2, column=j + 1, value=value)

# Guarda los cambios en el libro de trabajo
book.save(ruta_excel)